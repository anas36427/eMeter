
# Django core imports
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.middleware.csrf import get_token
from django.utils import timezone
from django.db import models
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db import transaction, OperationalError
from django.conf import settings
from django.core.exceptions import ValidationError

# Standard library
from datetime import datetime, timedelta
import json
import random
import string
import io
import os

# Internal app imports
from electricity_system.authentication import require_authenticated, require_role
from billing.pdf_generator import BillPDFGenerator
from billing.services import BillingService
from billing.models import Consumer, MeterReading, Bill, Payment, BillingSettings, AuditLog
from billing.forms import ConsumerRegistrationForm

# PDF generation
from reportlab.lib import colors  # ← FIXED: Added missing import
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Django REST Framework
from rest_framework.decorators import api_view, authentication_classes
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.authentication import SessionAuthentication

# Use the custom User model — never import User directly
User = get_user_model()


def get_csrf(request):
    get_token(request)
    return JsonResponse({'status': 'ok'})

def api_root(request):
    """Health check endpoint for /api/"""
    return JsonResponse({'status': 'ok', 'message': 'API is running', 'version': '1.0'})

def favicon_view(request):
    """Return an empty response for favicon.ico to prevent 404 errors"""
    return HttpResponse(status=204)

def login_view(request):
    """Handle user login"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Determine role from database only
            role = getattr(user, 'role', 'meter_reader')
            if role == 'admin':
                return redirect('admin_dashboard')
            elif role == 'meter_reader':
                return redirect('meter_reader_dashboard')
            else:
                return redirect('consumer_dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'login.html')


def logout_view(request):
    """Handle user logout"""
    logout(request)
    return redirect('login')

@csrf_exempt
def api_login(request):
    """JSON login endpoint for SPA clients"""
    if settings.DEBUG:
        print(f"DEBUG: api_login hit with method {request.method}")
    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed'}, status=405)

    # --- LIGHTWEIGHT FAILED-LOGIN RATE LIMITING ---
    from django.core.cache import cache
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', '127.0.0.1')

    cache_key = f"login_attempts_{ip}"
    attempts = cache.get(cache_key, 0)

    if attempts >= 5:
        return JsonResponse({'detail': 'Too many failed login attempts. Please try again in 5 minutes.'}, status=429)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'detail': 'Invalid JSON body'}, status=400)

    username = data.get('username')
    password = data.get('password')
    if settings.DEBUG:
        print(f"DEBUG: Attempting login for {username}")

    if not username or not password:
        return JsonResponse({'detail': 'Username and password are required'}, status=400)

    user = authenticate(request, username=username, password=password)

    if user is None:
        if settings.DEBUG:
            print(f"DEBUG: Authentication failed for {username}")
        # Increment failed login attempt count
        cache.set(cache_key, attempts + 1, timeout=300)
        return JsonResponse({'detail': 'Invalid username or password.'}, status=401)

    # BUG-20 FIX: Prevent admins from logging in via the mobile app
    if getattr(user, 'role', 'meter_reader') == 'admin' and data.get('source') == 'mobile':
        return JsonResponse({
            'detail': 'Administrators cannot log in via the mobile app. Please use the Web Dashboard.'
        }, status=403)

    # Success: clear failed attempts
    cache.delete(cache_key)

    login(request, user)
    request.session.save()
    csrf_token = get_token(request)
    
    # LOGIC-04 FIX: Rotate token on every login for security
    Token.objects.filter(user=user).delete()
    token = Token.objects.create(user=user)
    if settings.DEBUG:
        print(f"DEBUG: Login successful for {user.username}. Role: {user.role}")
    return JsonResponse({
        'success': True,
        'username': user.username,
        'role': user.role,
        'token': token.key,
        'csrftoken': csrf_token,
    })

@api_view(['GET'])
def api_me(request):
    if not request.user.is_authenticated:
        return JsonResponse({'detail': 'Not authenticated.'}, status=401)
    return JsonResponse({
        'authenticated': True,
        'username': request.user.username,
        'role': request.user.role,
        'first_name': request.user.first_name,
        'last_name': request.user.last_name,
        'email': request.user.email,
    })

@api_view(['POST'])
def api_update_profile(request):
    """Update current user profile info"""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    try:
        data = request.data
        user = request.user
        
        # Split first and last names if full name is provided
        if 'name' in data:
            parts = data['name'].strip().split(' ', 1)
            user.first_name = parts[0]
            user.last_name = parts[1] if len(parts) > 1 else ''
        elif 'first_name' in data or 'last_name' in data:
            if 'first_name' in data:
                user.first_name = data['first_name']
            if 'last_name' in data:
                user.last_name = data['last_name']
                
        if 'email' in data:
            user.email = data['email']
            
        user.save()
        return JsonResponse({'success': True, 'message': 'Profile updated successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# 1.
@api_view(['GET'])
@require_role('admin', 'meter_reader')
def api_dashboard_stats(request):
    """JSON stats for SPA dashboard"""
    from django.utils import timezone
    now = timezone.now()
    
    total_consumers = Consumer.objects.count()
    total_bills = Bill.objects.count()
    total_readings = MeterReading.objects.count()
    
    total_revenue = Bill.objects.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or 0
    # BUG-12 FIX: 'unpaid' is not a valid Bill status — the correct status for billed-but-unpaid is 'finalized'.
    pending_amount = Bill.objects.filter(status='finalized').aggregate(total=Sum('total_amount'))['total'] or 0
    
    current_month_units = Bill.objects.filter(
        billing_period_start__month=now.month, 
        billing_period_start__year=now.year
    ).aggregate(total=Sum('units'))['total'] or 0

    return JsonResponse({
        'total_consumers': total_consumers,
        'total_bills': total_bills,
        'total_readings': total_readings,
        'total_revenue': round(float(total_revenue), 2),
        'pending_amount': round(float(pending_amount), 2),
        'current_month_units': round(float(current_month_units), 1),
    })


@api_view(['GET'])
@require_role('admin')
def api_reports_data(request):
    """API endpoint for reports charts data"""
    
    try:
        # Top 10 Consumers by Usage (Units)
        top_consumers = Bill.objects.values('consumer__name').annotate(
            total_units=Sum('units')
        ).order_by('-total_units')[:10]
        
        top_consumers_data = [
            {'name': item['consumer__name'] or "Unknown", 'units': round(float(item['total_units'] or 0), 1)}
            for item in top_consumers
        ]
        
        # Monthly Usage (last 6 months)
        monthly_usage = Bill.objects.values('billing_period_start').annotate(
            total_units=Sum('units')
        ).order_by('-billing_period_start')[:6]
        
        monthly_usage_data = [
            {'month': item['billing_period_start'].strftime('%b') if item['billing_period_start'] else "N/A", 'units': round(float(item['total_units'] or 0), 1)}
            for item in monthly_usage
        ]
        monthly_usage_data.reverse()
        
        # Revenue Breakdown (Salary vs Non-Salary)
        revenue_breakdown = Bill.objects.filter(status='paid').values('consumer__connection_type').annotate(
            total_revenue=Sum('total_amount')
        )
        
        revenue_data = []
        for item in revenue_breakdown:
            label = 'Salary' if item['consumer__connection_type'] == 'salary' else 'Non Salary'
            revenue_data.append({'name': label, 'value': round(float(item['total_revenue'] or 0), 2)})
            
        return JsonResponse({
            'top_consumers': top_consumers_data,
            'monthly_usage': monthly_usage_data,
            'revenue_breakdown': revenue_data,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def dashboard(request):
    """Main dashboard view - redirects based on role"""
    # ← FIXED: Using user.role directly instead of user.profile.role
    if request.user.role == 'admin':
        return redirect('admin_dashboard')
    elif request.user.role == 'meter_reader':
        return redirect('meter_reader_dashboard')
    else:
        return redirect('consumer_dashboard')


@login_required
def admin_dashboard(request):
    """Admin dashboard with stats and management"""
    # Get statistics
    total_consumers = Consumer.objects.count()
    active_consumers = Consumer.objects.filter(status='active').count()
    total_bills = Bill.objects.count()
    paid_bills = Bill.objects.filter(status='paid').count()
    # BUG-13 FIX: 'unpaid' and 'overdue' are not valid Bill statuses.
    # Finalized = bills that have been locked/issued but not yet paid.
    # Overdue = finalized bills whose due date has already passed.
    unpaid_bills = Bill.objects.filter(status='finalized').count()
    overdue_bills = Bill.objects.filter(
        status='finalized',
        due_date__lt=timezone.now().date()
    ).count()
    
    # Revenue calculation
    total_revenue = Bill.objects.filter(status='paid').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    # Recent readings
    recent_readings = MeterReading.objects.select_related('consumer').order_by('-created_at')[:5]
    
    # Recent consumers
    recent_consumers = Consumer.objects.order_by('-created_at')[:5]
    
    # All active consumers for the Generate Bill dropdown
    consumers = Consumer.objects.filter(status='active')
    
    context = {
        'total_consumers': total_consumers,
        'active_consumers': active_consumers,
        'total_bills': total_bills,
        'paid_bills': paid_bills,
        'unpaid_bills': unpaid_bills,
        'overdue_bills': overdue_bills,
        'total_revenue': total_revenue,
        'recent_readings': recent_readings,
        'recent_consumers': recent_consumers,
        'consumers': consumers,
    }
    
    return render(request, 'admin-dashboard.html', context)


@login_required
def manage_consumers(request):
    """Manage all consumers"""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    consumers = Consumer.objects.all().order_by('-created_at')
    
    if search_query:
        consumers = consumers.filter(
            Q(name__icontains=search_query) |
            Q(consumer_number__icontains=search_query) |
            Q(meter_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if status_filter:
        consumers = consumers.filter(status=status_filter)
    
    context = {
        'consumers': consumers,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    # If HTMX requests the partial table, render only the list
    if request.headers.get('HX-Request') or request.headers.get('Hx-Request') or request.META.get('HTTP_HX_REQUEST'):
        return render(request, 'partials/consumers_list.html', {'consumers': consumers})

    return render(request, 'manage-consumers.html', context)

@login_required
def add_consumer(request):
    """Add new consumer"""
    if request.method == 'POST':
        form = ConsumerRegistrationForm(request.POST)  # BUG-44 FIX: was undefined ConsumerForm
        if form.is_valid():
            form.save()
            messages.success(request, 'Consumer added successfully!')
            return redirect('manage_consumers')
    else:
        form = ConsumerRegistrationForm()  # BUG-44 FIX
    
    return render(request, 'add-consumer.html', {'form': form})


@login_required
def generate_bill(request):
    if request.method == 'POST':
        consumer_id = request.POST.get('consumer')
        units = float(request.POST.get('units', 0))
        rate = float(request.POST.get('rate_per_unit', 7.50))
        fixed_charges = float(request.POST.get('fixed_charges', 125.00))
        billing_period_start_str = request.POST.get('billing_period_start')
        due_date = request.POST.get('due_date')

        consumer = get_object_or_404(Consumer, id=consumer_id)

        # Parse billing_period_start string (format: YYYY-MM) to date
        if billing_period_start_str:
            try:
                billing_period_start = datetime.strptime(billing_period_start_str + '-01', '%Y-%m-%d').date()
            except ValueError:
                billing_period_start = timezone.localtime().date()
        else:
            billing_period_start = timezone.localtime().date()

        # Parse due_date string to date
        if due_date:
            try:
                due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
            except ValueError:
                due_date = (timezone.localtime() + timedelta(days=15)).date()
        else:
            due_date = (timezone.localtime() + timedelta(days=15)).date()

        # Get previous reading
        last_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date').first()
        previous_reading = last_reading.current_reading if last_reading else 0
        current_reading = previous_reading + units

        # Create meter reading
        meter_reading = MeterReading.objects.create(
            consumer=consumer,
            previous_reading=previous_reading,
            current_reading=current_reading,
            reading_date=timezone.localtime().date(),
            units_consumed=units,
            created_by=request.user
        )

        # Create bill
        bill = Bill.objects.create(
            consumer=consumer,
            meter_reading=meter_reading,
            units=units,
            rate_per_unit=rate,
            fixed_charges=fixed_charges,
            billing_period_start=billing_period_start,
            due_date=due_date
        )

        # BUG-15 FIX: Calculate all charges using the service so total_amount is never 0.
        BillingService.calculate_bill(bill.id)
        bill.refresh_from_db()


        if request.headers.get('HX-Request') or request.headers.get('Hx-Request') or request.META.get('HTTP_HX_REQUEST'):
            html = render_to_string('partials/generate_bill_response.html', {'bill': bill}, request=request)
            return HttpResponse(html)
        # ── Generate PDF using new AMU eMeter Generator ──
        try:
            pdf_data = {
                'bill_number': bill.bill_number,
                'bill_date': bill.created_at.strftime('%d %b %Y') if bill.created_at else timezone.localtime().strftime('%d %b %Y'),
                'due_date': bill.due_date.strftime('%d %b %Y') if bill.due_date else 'N/A',
                'connection_type': bill.consumer.connection_type,
                'load_kw': bill.consumer.load_kw,
                'billing_period_start': bill.billing_period_start.strftime('%B %Y') if bill.billing_period_start else 'N/A',
                'consumer_name': bill.consumer.name,
                'consumer_number': bill.consumer.consumer_number,
                'meter_number': bill.consumer.meter_number,
                'address': bill.consumer.address,
                'previous_reading': bill.meter_reading.previous_reading if bill.meter_reading else 0,
                'current_reading': bill.meter_reading.current_reading if bill.meter_reading else 0,
                'units': bill.units,
                'rate_per_unit': bill.rate_per_unit,
                'energy_charges': bill.energy_charges,
                'fixed_charges': bill.fixed_charges,
                'duty_charge': bill.duty_charge,
                'meter_rent': bill.meter_rent,
                'meter_type': '1' if bill.consumer.meter_type == 'analog' else '3',
                'regulatory_surcharge': bill.regulatory_surcharge,
                'arrears': bill.arrears,
                'late_payment_surcharge': bill.late_payment_surcharge,
                'total_amount': bill.total_amount,
                'total_payable': int(round(bill.total_amount)),
                'current_year': timezone.localtime().year,
            }
            
            pdf_content = BillPDFGenerator.generate_bill_pdf(pdf_data)
            
            if pdf_content:
                response = HttpResponse(pdf_content, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="AMU_Bill_{bill.bill_number}.pdf"'
                return response
            else:
                return HttpResponse("Error generating PDF", status=500)
        except Exception as e:
            return HttpResponse(f"PDF Generation Failed: {str(e)}", status=500)

    consumers = Consumer.objects.filter(status='active')
    return render(request, 'generate-bill.html', {'consumers': consumers})

@login_required
def meter_readings(request):
    """View all meter readings"""
    readings = MeterReading.objects.select_related('consumer').order_by('-reading_date')
    return render(request, 'meter-readings.html', {'readings': readings})


@login_required
def bills(request):
    """View all bills"""
    status_filter = request.GET.get('status', '')
    
    bills = Bill.objects.select_related('consumer').order_by('-created_at')
    
    if status_filter:
        bills = bills.filter(status=status_filter)

    # HTMX partial support
    if request.headers.get('HX-Request') or request.headers.get('Hx-Request') or request.META.get('HTTP_HX_REQUEST'):
        return render(request, 'partials/bills_list.html', {'bills': bills})

    return render(request, 'bills.html', {'bills': bills, 'status_filter': status_filter})


@login_required
def payments(request):
    """View all payments"""
    payments_list = Payment.objects.select_related('bill__consumer').order_by('-payment_date')
    return render(request, 'payments.html', {'payments': payments_list})


# Consumer Portal Views
@login_required
def consumer_dashboard(request):
    """Consumer dashboard view"""
    try:
        consumer = request.user.consumer
    except Consumer.DoesNotExist:
        # Try to find consumer by email
        consumer = Consumer.objects.filter(email=request.user.email).first()
        if consumer:
            consumer.user = request.user
            consumer.save()
        else:
            messages.error(request, 'No consumer account found for this user')
            return redirect('login')
    
    # Get current bill
    current_bill = Bill.objects.filter(consumer=consumer, status='unpaid').first()
    
    # Get payment history
    payment_history = Payment.objects.filter(
        bill__consumer=consumer
    ).select_related('bill').order_by('-payment_date')[:5]
    
    # Get recent readings
    recent_readings = MeterReading.objects.filter(
        consumer=consumer
    ).order_by('-reading_date')[:6]
    
    # Calculate usage stats
    this_month_reading = recent_readings.first()
    last_month_reading = recent_readings[1] if len(recent_readings) > 1 else None
    
    usage_comparison = 0
    if this_month_reading and last_month_reading:
        if last_month_reading.units_consumed > 0:
            usage_comparison = ((this_month_reading.units_consumed - last_month_reading.units_consumed) 
                              / last_month_reading.units_consumed * 100)
    
    context = {
        'consumer': consumer,
        'current_bill': current_bill,
        'payment_history': payment_history,
        'recent_readings': recent_readings,
        'usage_comparison': usage_comparison,
    }
    
    return render(request, 'consumer-portal.html', context)


@login_required
def consumer_bills(request):
    """View all bills for current consumer"""
    try:
        consumer = request.user.consumer
    except Consumer.DoesNotExist:
        consumer = Consumer.objects.filter(email=request.user.email).first()
        if consumer:
            consumer.user = request.user
            consumer.save()
        else:
            messages.error(request, 'No consumer account found for this user')
            return redirect('login')
    
    bills = Bill.objects.filter(consumer=consumer).order_by('-billing_period_start')
    return render(request, 'consumer-bills.html', {'bills': bills})


@login_required
def bill_details_partial(request, bill_id):
    """Return bill details fragment for HTMX"""
    bill = get_object_or_404(Bill, id=bill_id)
    return render(request, 'partials/consumer_bill_details.html', {'bill': bill})


@login_required
def make_payment(request, bill_id):
    """Process payment for a bill"""
    bill = get_object_or_404(Bill, id=bill_id, consumer__user=request.user)
    
    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')
        
        # Generate transaction ID
        transaction_id = 'TXN' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
        
        # Create payment
        Payment.objects.create(
            bill=bill,
            transaction_id=transaction_id,
            amount=bill.total_amount,
            payment_method=payment_method,
            status='success'
        )
        
        # Update bill status
        bill.status = 'paid'
        bill.paid_date = timezone.localtime().date()
        bill.transaction_id = transaction_id
        bill.save()
        
        messages.success(request, f'Payment successful! Transaction ID: {transaction_id}')
        return redirect('consumer_dashboard')
    
    return render(request, 'make-payment.html', {'bill': bill})


# Meter Reader Views
@login_required
def meter_reader_dashboard(request):
    """Meter reader dashboard"""
    # ← FIXED: Using user.role directly instead of user.profile.role
    if request.user.role not in ['meter_reader', 'admin']:
        return redirect('dashboard')
    
    # Get assigned consumers (for now, all active consumers)
    assigned_consumers = Consumer.objects.filter(status='active')
    
    # Get today's readings count
    today = timezone.localtime().date()
    today_readings = MeterReading.objects.filter(reading_date=today).count()
    
    # Get pending readings (consumers without readings this month)
    this_month = today.replace(day=1)
    consumers_with_readings = MeterReading.objects.filter(
        reading_date__gte=this_month
    ).values_list('consumer_id', flat=True)
    
    pending_readings = assigned_consumers.exclude(id__in=consumers_with_readings).count()
    
    context = {
        'assigned_consumers': assigned_consumers,
        'total_assigned': assigned_consumers.count(),
        'today_readings': today_readings,
        'pending_readings': pending_readings,
        'completion_rate': round((today_readings / assigned_consumers.count() * 100) if assigned_consumers.count() > 0 else 0, 1),
    }
    
    return render(request, 'meter-reader.html', context)


@login_required
def submit_reading(request):
    """Submit meter reading"""
    if request.method == 'POST':
        consumer_id = request.POST.get('consumer')
        previous_reading = float(request.POST.get('previous_reading', 0))
        current_reading = float(request.POST.get('current_reading', 0))
        reading_date = request.POST.get('reading_date')
        reading_time = request.POST.get('reading_time')
        remarks = request.POST.get('remarks', '')
        
        if current_reading < previous_reading:
            messages.error(request, 'Current reading must be greater than or equal to previous reading')
            return redirect('submit_reading')
        
        consumer = get_object_or_404(Consumer, id=consumer_id)
        
        # Parse time if provided
        reading_time_obj = None
        if reading_time:
            try:
                reading_time_obj = datetime.strptime(reading_time, '%H:%M').time()
            except ValueError:
                pass
        
        # Create meter reading
        reading = MeterReading.objects.create(
            consumer=consumer,
            previous_reading=previous_reading,
            current_reading=current_reading,
            reading_date=reading_date,
            reading_time=reading_time_obj,
            remarks=remarks,
            created_by=request.user
        )

        # If this is an HTMX request, return a small fragment that HTMX will swap
        if request.headers.get('HX-Request') or request.headers.get('Hx-Request') or request.META.get('HTTP_HX_REQUEST'):
            units = (current_reading - previous_reading) if current_reading >= previous_reading else 0
            context = {
                'consumer_id': consumer.id,
                'previous_reading': current_reading,
                'current_reading': current_reading,
                'units_consumed': units,
            }
            html = render_to_string('partials/reading_response.html', context, request=request)
            return HttpResponse(html)

        messages.success(request, 'Reading submitted successfully!')
        return redirect('meter_reader_dashboard')
    
    # Get consumers with their last reading
    consumers = Consumer.objects.filter(status='active')
    consumer_data = []
    
    for consumer in consumers:
        last_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date').first()
        consumer_data.append({
            'consumer': consumer,
            'previous_reading': last_reading.current_reading if last_reading else 0
        })
    
    today = timezone.localtime().date().strftime('%Y-%m-%d')
    
    return render(request, 'submit-reading.html', {
        'consumer_data': consumer_data,
        'today': today
    })


# API Views for AJAX calls

# .1
@require_authenticated
def api_get_consumer(request, consumer_id):
    """Get consumer details for API"""
    consumer = get_object_or_404(Consumer, id=consumer_id)
    
    # Check authorization to prevent BOLA/IDOR
    if request.user.role not in ['admin', 'meter_reader']:
        is_owner = False
        if hasattr(request.user, 'consumer_profile') and request.user.consumer_profile and request.user.consumer_profile.id == consumer.id:
            is_owner = True
        elif hasattr(request.user, 'consumer') and request.user.consumer and request.user.consumer.id == consumer.id:
            is_owner = True
        elif consumer.user == request.user:
            is_owner = True
            
        if not is_owner:
            return JsonResponse({'detail': 'Access denied.'}, status=403)

    last_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date').first()
    
    data = {
        'id': consumer.id,
        'name': consumer.name,
        'meter_number': consumer.meter_number,
        'consumer_number': consumer.consumer_number,
        'address': consumer.address,
        'status': consumer.status,
        'previous_reading': last_reading.current_reading if last_reading else 0,
        # Safe fallbacks from model properties/methods
        'load_kw': getattr(consumer, 'load_kw', 1.0),
        'meter_type': getattr(consumer, 'meter_type', '10'),
    }
    
    return JsonResponse(data)

# .3
@require_authenticated
def api_calculate_bill(request):
    """Calculate bill amount"""
    if request.method == 'GET':
        units = float(request.GET.get('units', 0))
        rate = float(request.GET.get('rate', 7.50))
        fixed_charges = float(request.GET.get('fixed_charges', 125.00))
        
        energy_charges = units * rate
        total = energy_charges + fixed_charges
        
        return JsonResponse({
            'units': units,
            'energy_charges': round(energy_charges, 2),
            'fixed_charges': fixed_charges,
            'total_amount': round(total, 2)
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

# 4.
@api_view(['GET', 'POST'])
@require_role('admin', 'meter_reader')
def api_consumer_list(request):

    if request.method == 'GET':
        consumers = Consumer.objects.all()
        results = []
        for c in consumers:
            # Get the most recent reading for this consumer
            last_reading = MeterReading.objects.filter(consumer=c).order_by('-reading_date', '-id').first()
            prev_val = last_reading.current_reading if last_reading else 0
            
            results.append({
                'id': c.id,
                'name': c.name,
                'meter_number': c.meter_number,
                'consumer_number': c.consumer_number,
                'address': c.address,
                'status': c.status,
                'load_kw': c.load_kw,
                'meter_type': c.meter_type,
                'connection_type': c.connection_type,
                'billing_type': c.billing_type,
                'previous_reading': prev_val
            })
        return JsonResponse({'consumers': results})

    if request.method == 'POST':
        try:
            data = request.data
            # Validate required fields
            if not data.get('name') or not data.get('meter_number'):
                return JsonResponse({'success': False, 'error': 'Name and Meter Number are required.'}, status=400)
            # Check for duplicate meter number
            if Consumer.objects.filter(meter_number=data['meter_number']).exists():
                return JsonResponse({'success': False, 'error': 'A consumer with this meter number already exists.'}, status=400)

            # Manual consumer ID or auto-generate
            consumer_number = data.get('consumer_number', '').strip()
            if consumer_number:
                if Consumer.objects.filter(consumer_number=consumer_number).exists():
                    return JsonResponse({'success': False, 'error': 'A consumer with this ID already exists.'}, status=400)
            else:
                consumer_number = None
                for _ in range(10):
                    candidate = 'CN' + ''.join(random.choices(string.digits, k=6))
                    if not Consumer.objects.filter(consumer_number=candidate).exists():
                        consumer_number = candidate
                        break
                if not consumer_number:
                    consumer_number = 'CN' + ''.join(random.choices(string.digits, k=10))

            # Normalize connection_type: handle legacy frontend values
            connection_type_raw = data.get('connection_type', 'single_phase')
            CONNECTION_TYPE_MAP = {
                'salary': 'single_phase',
                'non-salary': 'three_phase',
                'residential': 'single_phase',
                'commercial': 'three_phase',
                'single_phase': 'single_phase',
                'three_phase': 'three_phase',
            }
            connection_type = CONNECTION_TYPE_MAP.get(connection_type_raw, 'single_phase')

            # Normalize meter_type: handle legacy numeric frontend values
            meter_type_raw = data.get('meter_type', 'analog')
            METER_TYPE_MAP = {
                '10': 'analog',
                '25': 'digital',
                'analog': 'analog',
                'digital': 'digital',
                'smart': 'smart',
            }
            meter_type = METER_TYPE_MAP.get(str(meter_type_raw), 'analog')

            # Normalize billing_type
            billing_type_raw = data.get('billing_type', 'salary')
            BILLING_TYPE_MAP = {
                'salary': 'salary',
                'non-salary': 'non_salary',
                'non_salary': 'non_salary',
            }
            billing_type = BILLING_TYPE_MAP.get(billing_type_raw, 'salary')

            consumer = Consumer.objects.create(
                name=data.get('name'),
                phone=data.get('phone', ''),
                email=data.get('email') or None,
                address=data.get('address', ''),
                post=data.get('post', ''),
                department=data.get('department', ''),
                meter_number=data.get('meter_number'),
                load_kw=float(data.get('load_kw', 1.0)),
                meter_type=meter_type,
                connection_type=connection_type,
                billing_type=billing_type,
                status=data.get('status', 'active'),
                consumer_number=consumer_number,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )

            # Auto-generate portal account for all new consumers
            try:
                user = User.objects.create_user(
                    username=consumer.consumer_number,
                    password=consumer.consumer_number,
                    role='consumer'
                )
                consumer.user = user
                consumer.save()
            except Exception as e:
                print(f"Failed to auto-create portal account for {consumer.consumer_number}: {e}")

            # Create initial meter reading if provided
            initial_reading = data.get('initial_reading')
            if initial_reading is not None:
                try:
                    reading_val = float(initial_reading)
                    MeterReading.objects.create(
                        consumer=consumer,
                        previous_reading=0,
                        current_reading=reading_val,
                        reading_date=timezone.now().date(),
                        reading_time=timezone.now().time(),
                        remarks='Initial reading at registration',
                        created_by=request.user if request.user.is_authenticated else None,
                        created_at=timezone.now(),
                    )
                except (ValueError, TypeError):
                    pass  # Skip if invalid number

            return JsonResponse({
                'success': True,
                'id': consumer.id,
                'consumer_number': consumer.consumer_number,
                'name': consumer.name,
                'meter_number': consumer.meter_number,
            }, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'detail': 'Method not allowed'}, status=405)

# 5.
@api_view(['GET'])
@require_role('admin', 'meter_reader')
def api_bills_list(request):
    """Get all bills for admin with Server-Side Pagination and Search"""
    if request.method == 'GET':
        page_num = request.GET.get('page', 1)
        limit = request.GET.get('limit', 50)
        search_query = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', 'all')
        account_type = request.GET.get('accountType', 'all')

        sort_order = request.GET.get('sortOrder', 'desc')
        start_date = request.GET.get('startDate', '')
        end_date = request.GET.get('endDate', '')

        bills = Bill.objects.select_related('consumer').all()

        # Apply Server-Side Search
        if search_query:
            bills = bills.filter(
                Q(bill_number__icontains=search_query) |
                Q(consumer__name__icontains=search_query) |
                Q(consumer__consumer_number__icontains=search_query) |
                Q(consumer__meter_number__icontains=search_query)
            )

        # Apply Filters
        if status_filter != 'all':
            if status_filter.lower() == 'unpaid':
                bills = bills.filter(status='finalized')
            else:
                bills = bills.filter(status=status_filter.lower())
        
        if account_type == 'salary':
            bills = bills.filter(consumer__billing_type='salary')
        elif account_type == 'non-salary':
            bills = bills.filter(consumer__billing_type='non_salary')
            
        if start_date:
            try:
                # Format: YYYY-MM
                year, month = start_date.split('-')
                bills = bills.filter(billing_period_start__gte=datetime(int(year), int(month), 1))
            except Exception:
                pass
                
        if end_date:
            try:
                # Format: YYYY-MM
                year, month = end_date.split('-')
                if int(month) == 12:
                    next_month = datetime(int(year) + 1, 1, 1)
                else:
                    next_month = datetime(int(year), int(month) + 1, 1)
                bills = bills.filter(billing_period_start__lt=next_month)
            except Exception:
                pass

        # Apply Sort
        if sort_order == 'asc':
            bills = bills.order_by('created_at')
        else:
            bills = bills.order_by('-created_at')

        # Calculate totals for the entire filtered queryset
        totals = bills.aggregate(
            total_billed=Sum('total_amount'),
            total_paid=Sum('total_amount', filter=Q(status='paid'))
        )
        total_billed = float(totals['total_billed'] or 0)
        total_paid = float(totals['total_paid'] or 0)
        total_pending = total_billed - total_paid

        # Pagination
        paginator = Paginator(bills, int(limit))
        try:
            page_obj = paginator.page(page_num)
        except Exception:
            page_obj = paginator.page(1)

        results = []
        for b in page_obj.object_list:
            results.append({
                'id': b.id,
                'bill_number': b.bill_number,
                'consumer_name': b.consumer.name if b.consumer else 'N/A',
                'consumer_number': b.consumer.consumer_number if b.consumer else 'N/A',
                'meter_number': b.consumer.meter_number if b.consumer else 'N/A',
                'units': b.units,
                'total_amount': b.total_amount,
                'status': b.status,
                'billing_period_start': b.billing_period_start.strftime('%B %Y') if b.billing_period_start else '',
                'connection_type': b.consumer.connection_type if b.consumer else 'residential',
                'billing_type': b.consumer.billing_type if b.consumer else 'non_salary',
                'due_date': b.due_date.strftime('%Y-%m-%d') if b.due_date else '',
                'created_at': b.created_at.isoformat() if b.created_at else None,
            })
            
        return JsonResponse({
            'bills': results,
            'total_items': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'summary': {
                'total_billed': total_billed,
                'total_paid': total_paid,
                'total_pending': total_pending
            }
        })

@require_authenticated
def api_logout(request):
    """JSON logout endpoint for SPA clients"""
    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed'}, status=405)
        
    # LOGIC-04 FIX: Destroy token on logout so it can't be reused by a stolen device
    if request.user.is_authenticated:
        Token.objects.filter(user=request.user).delete()
        
    logout(request)
    return JsonResponse({'success': True, 'message': 'Logged out successfully'})

# 7.
@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@require_role('admin', 'meter_reader')
def api_consumer_detail(request, consumer_id):
    """Get, update, or delete a consumer"""
    consumer = get_object_or_404(Consumer, id=consumer_id)

    if request.method == 'GET':
        # Get last reading
        last_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date', '-id').first()
        prev_val = last_reading.current_reading if last_reading else 0

        data = {
            'id': consumer.id,
            'name': consumer.name,
            'meter_number': consumer.meter_number,
            'consumer_number': consumer.consumer_number,
            'email': consumer.email,
            'phone': consumer.phone,
            'address': consumer.address,
            'status': consumer.status,
            'load_kw': str(consumer.load_kw) if consumer.load_kw else '',
            'meter_type': consumer.meter_type,
            'connection_type': consumer.connection_type,
            'department': consumer.department,
            'post': consumer.post,
            'billing_type': consumer.billing_type,
            'previous_reading': prev_val,
        }
        return JsonResponse(data)

    if request.method in ['PUT', 'PATCH']:
        try:
            data = request.data
            for key, value in data.items():
                if hasattr(consumer, key):
                    # Handle empty strings for numeric fields
                    if key == 'load_kw':
                        try:
                            value = float(value) if value else 1.0
                        except ValueError:
                            value = 1.0
                    elif value == "":
                        try:
                            field = consumer._meta.get_field(key)
                            if isinstance(field, (models.FloatField, models.IntegerField)) and not field.null:
                                continue
                        except Exception:
                            pass
                    
                    # Prevent overwriting with empty consumer number
                    if key == 'consumer_number' and not value:
                        continue
                        
                    setattr(consumer, key, value)
            consumer.save()
            return JsonResponse({'success': True, 'message': 'Consumer updated successfully'})
        except Exception as e:
            import traceback
            print(f"PATCH Error: {str(e)}\n{traceback.format_exc()}")
            return JsonResponse({'success': False, 'error': f"Server error: {str(e)}"}, status=400)

    if request.method == 'DELETE':
        consumer.delete()
        return JsonResponse({'success': True, 'message': 'Consumer deleted successfully'})

@api_view(['GET'])
@require_authenticated
def api_consumer_readings(request, consumer_id):
    """Get reading history for a specific consumer"""
    
    # Check authorization to prevent BOLA/IDOR
    if request.user.role not in ['admin', 'meter_reader']:
        is_owner = False
        if hasattr(request.user, 'consumer_profile') and request.user.consumer_profile and request.user.consumer_profile.id == consumer_id:
            is_owner = True
        elif hasattr(request.user, 'consumer') and request.user.consumer and request.user.consumer.id == consumer_id:
            is_owner = True
        else:
            is_owner = Consumer.objects.filter(id=consumer_id, user=request.user).exists()
            
        if not is_owner:
            return JsonResponse({'detail': 'Access denied.'}, status=403)

    readings = MeterReading.objects.filter(consumer_id=consumer_id).order_by('-reading_date')
    results = []
    for reading in readings:
        results.append({
            'id': reading.id,
            'date': reading.reading_date.strftime('%Y-%m-%d') if reading.reading_date else '',
            'reading': reading.current_reading,
            'prev': reading.previous_reading,
            'usage': reading.units_consumed,
            'recorded_by': reading.created_by.username if reading.created_by else 'Admin',
            'remarks': reading.remarks or '',
            'created_at': reading.created_at.isoformat() if reading.created_at else None,
        })
    return JsonResponse(results, safe=False)


# 8.
@api_view(['GET'])
@require_authenticated
def api_consumer_search(request):
    """Search consumers by meter number, name, or ID"""
    query = request.GET.get('meter_number', '')  # Param name remains for compatibility
    if not query:
        return JsonResponse({'detail': 'search parameter required'}, status=400)
    
    # BUG-22 FIX: id is an integer field. Using icontains on it causes a crash on PostgreSQL.
    # We dynamically construct the filter, using exact match `id=int(query)` only if the query is numeric.
    search_filter = (
        Q(meter_number__icontains=query) |
        Q(name__icontains=query) |
        Q(consumer_number__icontains=query)
    )
    if query.isdigit():
        search_filter |= Q(id=int(query))

    # Check authorization to prevent BOLA/IDOR
    if request.user.role not in ['admin', 'meter_reader']:
        user_consumer = None
        if hasattr(request.user, 'consumer_profile') and request.user.consumer_profile:
            user_consumer = request.user.consumer_profile
        elif hasattr(request.user, 'consumer') and request.user.consumer:
            user_consumer = request.user.consumer
        else:
            user_consumer = Consumer.objects.filter(user=request.user).first()
            
        if not user_consumer:
            return JsonResponse({'consumers': []})
            
        search_filter = Q(id=user_consumer.id) & search_filter

    try:
        consumers = Consumer.objects.filter(search_filter)
        # Trigger query to check for missing columns
        list(consumers[:1]) 
    except Exception as e:
        err_msg = str(e).lower()
        if 'no such column' in err_msg or 'does not exist' in err_msg:
            consumers = consumers.defer('load_kw', 'meter_type')
        else:
            raise e

    results = []
    for c in consumers:
        # Get last reading to help UI
        last_reading = MeterReading.objects.filter(consumer=c).order_by('-reading_date').first()
        results.append({
            'id': c.id,
            'name': c.name,
            'meter_number': c.meter_number,
            'consumer_number': c.consumer_number,
            'status': c.status,
            'previous_reading': last_reading.current_reading if last_reading else 0,
            'load_kw': getattr(c, 'load_kw', 1.0),
            'meter_type': getattr(c, 'meter_type', '10'),
            'address': c.address
        })
    return JsonResponse({'consumers': results})


#9
@api_view(['GET', 'POST'])
@require_role('admin', 'meter_reader')
def api_readings_list(request):
    """Get all meter readings (GET) or submit a new one (POST)"""
    
    if request.method == 'GET':
        readings = MeterReading.objects.select_related('consumer').order_by('-reading_date', '-id')
        
        # ── 1. Apply Month & Year Filters ──
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        if month and month.isdigit():
            readings = readings.filter(reading_date__month=int(month))
        if year and year.isdigit():
            readings = readings.filter(reading_date__year=int(year))
            
        total_count = readings.count()
        
        # ── 2. Handle Pagination if requested ──
        page = request.GET.get('page')
        limit = request.GET.get('limit', '10')
        
        has_more = False
        if page and page.isdigit():
            try:
                page_num = int(page)
                limit_num = int(limit) if limit.isdigit() else 10
                paginator = Paginator(readings, limit_num)
                page_obj = paginator.get_page(page_num)
                readings = page_obj.object_list
                has_more = page_obj.has_next()
            except Exception as e:
                print("Pagination error:", e)
        
        results = []
        for reading in readings:
            results.append({
                'id': reading.id,
                'consumer_id': reading.consumer.id,
                'consumer_name': reading.consumer.name,
                'consumer_number': reading.consumer.consumer_number,  # Add consumer number for excel export mapping
                'meter_number': reading.consumer.meter_number,
                'previous_reading': reading.previous_reading,
                'current_reading': reading.current_reading,
                'units_consumed': reading.units_consumed,
                'reading_date': reading.reading_date.strftime('%Y-%m-%d') if reading.reading_date else None,
                'created_at': reading.created_at.strftime('%Y-%m-%d %H:%M') if reading.created_at else None,
            })
            
        response_data = {'readings': results}
        if page:
            response_data.update({
                'page': int(page),
                'has_more': has_more,
                'total_count': total_count
            })
            
        return JsonResponse(response_data)
    
    if request.method == 'POST':
        return _do_submit_reading(request)

#10
@api_view(['GET', 'PATCH'])
@require_role('admin', 'meter_reader')
def api_bill_detail(request, bill_id):
    """Get or update a bill"""
    bill = get_object_or_404(Bill, id=bill_id)

    if request.method == 'GET':
        data = {
            'id': bill.id,
            'bill_number': bill.bill_number,
            'consumer_id': bill.consumer.id,
            'consumer_name': bill.consumer.name,
            'consumer_number': bill.consumer.consumer_number,
            'meter_number': bill.consumer.meter_number,
            'address': bill.consumer.address,
            'connection_type': bill.consumer.connection_type,
            'load_kw': bill.consumer.load_kw,
            'meter_type': bill.consumer.meter_type,
            'billing_type': bill.consumer.billing_type,
            'previous_reading': bill.meter_reading.previous_reading if bill.meter_reading else 0,
            'current_reading': bill.meter_reading.current_reading if bill.meter_reading else 0,
            'units': bill.units,
            'rate_per_unit': bill.rate_per_unit,
            'fixed_charges': bill.fixed_charges,
            'energy_charges': bill.energy_charges,
            'duty_charge': bill.duty_charge,
            'meter_rent': bill.meter_rent,
            'regulatory_surcharge': bill.regulatory_surcharge,
            'arrears': bill.arrears,
            'late_payment_surcharge': bill.late_payment_surcharge,
            'total_amount': bill.total_amount,
            'status': bill.status,
            'billing_period_start': bill.billing_period_start.strftime('%B %Y') if bill.billing_period_start else None,
            'due_date': bill.due_date.strftime('%Y-%m-%d') if bill.due_date else None,
            'paid_date': bill.payment_date.strftime('%Y-%m-%d') if bill.payment_date else None,
            'created_at': bill.created_at.isoformat() if bill.created_at else None,
        }
        return JsonResponse(data)

    if request.method == 'PATCH':
        data = request.data
        if 'status' in data:
            new_status = data['status']
            # BUG-25 FIX: Locked (finalized) bills may only transition to 'paid' or 'cancelled'.
            # Allowing any other status would reopen a frozen bill, breaking audit integrity.
            if bill.is_locked and new_status not in ('paid', 'cancelled'):
                return JsonResponse(
                    {'error': f"Locked bills can only be marked as 'paid' or 'cancelled', not '{new_status}'."},
                    status=400
                )
            bill.status = new_status
            if new_status == 'paid':
                bill.payment_date = timezone.localtime().date()
            bill.save()
        return JsonResponse({'success': True, 'message': 'Bill updated successfully'})

    return JsonResponse({'detail': 'Method not allowed'}, status=405)

@api_view(['POST'])
@require_role('admin', 'meter_reader')
@transaction.atomic
def api_mark_bill_paid(request, bill_id):
    """Mark a finalized bill as paid"""
    try:
        bill = get_object_or_404(Bill.objects.select_for_update(), id=bill_id)
        if bill.status == 'draft':
            return JsonResponse({'success': False, 'error': 'Cannot pay a draft bill. Finalize it first.'}, status=400)
        
        bill.status = 'paid'
        bill.payment_date = timezone.now().date()
        bill.save()
        
        AuditLog.objects.create(
            user=request.user,
            action='payment_update',
            bill=bill,
            details={'new_status': 'paid'}
        )
        return JsonResponse({'success': True, 'message': 'Bill marked as paid'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@api_view(['POST'])
@require_role('admin', 'meter_reader')
@transaction.atomic
def api_mark_bill_unpaid(request, bill_id):
    """Revert a paid status (only for admins)"""
    # ← FIXED: Using user.role directly instead of user.profile.role
    if request.user.role != 'admin':
        return JsonResponse({'detail': 'Only admins can revert payment status'}, status=403)
        
    try:
        bill = get_object_or_404(Bill.objects.select_for_update(), id=bill_id)
        bill.status = 'issued'
        bill.payment_date = None
        bill.save()
        return JsonResponse({'success': True, 'message': 'Bill marked as unpaid'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def _do_submit_reading(request):
    """Decorator-free helper function to submit a meter reading safely without wrapper conflicts."""
    try:
        data = request.data
        consumer_id = data.get('consumer_id')
        current_reading = float(data.get('current_reading', 0))
        reading_date = data.get('reading_date')
        meter_id = data.get('meter_id')
        meter_number = data.get('meter_number')
        
        consumer = get_object_or_404(Consumer, id=consumer_id)
        
        from billing.models import Meter
        meter = None
        if meter_id:
            meter = Meter.objects.filter(id=meter_id).first()
        elif meter_number:
            meter = Meter.objects.filter(meter_number=meter_number).first()
            
        # Use BillingService for validation
        try:
            previous_reading = BillingService.validate_reading(consumer, reading_date, current_reading, meter=meter)
        except ValidationError as e:
            error_msg = e.message if hasattr(e, 'message') else str(e)
            return JsonResponse({'success': False, 'error': error_msg}, status=400)

        reading = MeterReading.objects.create(
            consumer=consumer,
            meter=meter,
            previous_reading=previous_reading,
            current_reading=current_reading,
            reading_date=reading_date,
            created_by=request.user
        )
        
        # Check for consumption anomaly
        check_consumption_anomaly(reading)

        # Create a sync/submission notification if it's from a meter reader or mobile device
        from billing.models import AdminNotification
        AdminNotification.objects.create(
            user=request.user,
            title="Mobile Sync Success",
            message=f"Reading for consumer {consumer.name} (Meter: {consumer.meter_number}) successfully synced and finalized.",
            notification_type="success",
            is_read=False
        )
        
        AuditLog.objects.create(
            user=request.user,
            action='reading_submit',
            details={'reading_id': reading.id, 'consumer_id': consumer.id}
        )

        return JsonResponse({
            'success': True,
            'reading_id': reading.id,
            'units_consumed': reading.units_consumed
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@api_view(['POST'])
@require_role('admin', 'meter_reader')
def api_submit_reading(request):
    """Simple reading submission (Admin/Reader)"""
    return _do_submit_reading(request)


_SPA_INDEX_CACHE = None

def spa_index(request, path=None):
    """Render the single-page application index from energy-hub-ui/build."""
    global _SPA_INDEX_CACHE
    
    if _SPA_INDEX_CACHE is not None:
        return HttpResponse(_SPA_INDEX_CACHE)

    import os
    build_dir = os.path.join(settings.BASE_DIR.parent, 'energy-hub-ui', 'dist')
    index_path = os.path.join(build_dir, 'index.html')
    
    try:
        with open(index_path, 'r', encoding='utf-8') as f:
            content = f.read()
            if not settings.DEBUG:
                _SPA_INDEX_CACHE = content
            return HttpResponse(content)
    except FileNotFoundError:
        return HttpResponse(
            f"Vite build not found at {index_path}. Please run 'npm run build' in energy-hub-ui.",
            status=404
        )


# ==========================================
# Mobile App API Endpoints
# ==========================================

#15 - Combined: Submit Reading + Generate Finalized Bill.
@api_view(['POST'])
@require_role('admin', 'meter_reader')
@transaction.atomic
def api_submit_reading_and_generate_bill(request):
    """
    Combined: Submit Reading + Generate Finalized Bill.
    This implements the 'Reading Submitted -> Bill Generated' part of the workflow.
    """
    try:
        data = request.data
        consumer_id = data.get('consumer_id')
        current_reading = float(data.get('current_reading', 0))
        reading_date = data.get('reading_date')

        if not consumer_id or not reading_date:
            return JsonResponse({'error': 'consumer_id and reading_date are required'}, status=400)

        meter_id = data.get('meter_id')
        meter_number = data.get('meter_number')
        
        consumer = get_object_or_404(Consumer, id=consumer_id)
        
        from billing.models import Meter
        meter = None
        if meter_id:
            meter = Meter.objects.filter(id=meter_id).first()
        elif meter_number:
            meter = Meter.objects.filter(meter_number=meter_number).first()

        # 1. Atomic Generation of Final Bill
        reading_date_obj = datetime.strptime(reading_date, '%Y-%m-%d').date()
        
        try:
            bill = BillingService.generate_final_bill(
                consumer=consumer,
                current_reading=current_reading,
                reading_date=reading_date_obj,
                user=request.user,
                meter=meter
            )
        except ValidationError as e:
            error_msg = e.message if hasattr(e, 'message') else str(e)
            already_exists = "already exists" in error_msg.lower()
            response_data = {'success': False, 'error': error_msg, 'already_exists': already_exists}
            if "cannot be less than previous" in error_msg.lower():
                response_data['previous_reading'] = error_msg
            return JsonResponse(response_data, status=400)

        # Check for consumption anomaly
        check_consumption_anomaly(bill.meter_reading)

        # Create a sync/submission notification if it's from a meter reader or mobile device
        from billing.models import AdminNotification
        AdminNotification.objects.create(
            user=request.user,
            title="Mobile Sync Success",
            message=f"Reading for consumer {consumer.name} (Meter: {consumer.meter_number}) successfully synced and finalized.",
            notification_type="success",
            is_read=False
        )

        return JsonResponse({
            'success': True,
            'message': 'Reading submitted and bill generated.',
            'bill_id': bill.id,
            'bill_number': bill.bill_number,
            'total_amount': bill.total_amount,
            'status': bill.status,
            'reading': {
                'id': bill.meter_reading.id,
                'current_reading': bill.meter_reading.current_reading,
                'previous_reading': bill.meter_reading.previous_reading,
                'units_consumed': bill.meter_reading.units_consumed,
            },
            'bill': {
                'id': bill.id,
                'bill_number': bill.bill_number,
                'total_amount': bill.total_amount,
                'units': bill.units,
                'status': bill.status,
                'energy_charges': bill.energy_charges,
                'fixed_charges': bill.fixed_charges,
                'grand_total': bill.total_amount,
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@api_view(['GET'])
@require_role('admin', 'meter_reader')
def api_get_settings(request):
    """Retrieve current billing settings"""
    
    settings_obj = BillingSettings.get_settings()
    data = {
        'rate_per_unit': settings_obj.rate_per_unit,
        'fixed_charge_per_kw': settings_obj.fixed_charge_per_kw,
        'phase_1_rent': settings_obj.phase_1_rent,
        'phase_3_rent': settings_obj.phase_3_rent,
        'duty_percentage': settings_obj.duty_percentage,
    }
    return JsonResponse(data)

@api_view(['POST'])
@require_role('admin')
def api_update_settings(request):
    """Update billing settings (admin only)"""
    
    try:
        data = request.data
        settings_obj = BillingSettings.get_settings()
        
        # Check if we got a mock object (no _state attribute is a good check for Django models)
        if not hasattr(settings_obj, '_state'):
            return JsonResponse({
                'success': False, 
                'error': 'Database error: Settings table is missing or database is locked. Please contact administrator.'
            }, status=500)

        if 'rate_per_unit' in data:
            settings_obj.rate_per_unit = float(data['rate_per_unit'])
        if 'fixed_charge_per_kw' in data:
            settings_obj.fixed_charge_per_kw = float(data['fixed_charge_per_kw'])
        if 'phase_1_rent' in data:
            settings_obj.phase_1_rent = float(data['phase_1_rent'])
        if 'phase_3_rent' in data:
            settings_obj.phase_3_rent = float(data['phase_3_rent'])
        if 'duty_percentage' in data:
            settings_obj.duty_percentage = float(data['duty_percentage'])
            
        settings_obj.save()
        return JsonResponse({'success': True, 'message': 'Settings updated successfully'})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': f'Invalid numeric value: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server Error: {str(e)}'}, status=500)


@api_view(['POST'])
@require_role('admin', 'meter_reader')
def api_calculate_estimate(request):
    """
    Calculate a real-time bill estimate using the live BillingSettings from the DB.
    This is the SINGLE SOURCE OF TRUTH for billing math — mirrors Bill.save() exactly.

    POST body:
        {
            "consumer_id": 5,
            "current_reading": 3450,
            "previous_reading": 3200   # optional — fetched from DB if omitted
        }

    Returns a full line-item breakdown so the mobile app never computes billing math locally.
    """

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON body'}, status=400)

    consumer_id = data.get('consumer_id')
    current_reading = data.get('current_reading')

    if consumer_id is None or current_reading is None:
        return JsonResponse({'error': 'consumer_id and current_reading are required'}, status=400)

    try:
        consumer = Consumer.objects.get(id=consumer_id)
    except Consumer.DoesNotExist:
        return JsonResponse({'error': 'Consumer not found'}, status=404)

    # Resolve previous reading from DB if not provided by the client
    if 'previous_reading' in data:
        previous_reading = float(data['previous_reading'])
    else:
        last_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date', '-id').first()
        previous_reading = last_reading.current_reading if last_reading else 0

    current_reading = float(current_reading)
    units_consumed = max(0, current_reading - previous_reading)

    # Fetch live settings from the database — no hardcoded fallbacks in the calculation
    billing_settings = BillingSettings.get_settings()
    rate_per_unit      = float(billing_settings.rate_per_unit)
    fixed_charge_per_kw = float(billing_settings.fixed_charge_per_kw)
    duty_percentage    = float(billing_settings.duty_percentage)
    phase_1_rent       = float(billing_settings.phase_1_rent)
    phase_3_rent       = float(billing_settings.phase_3_rent)

    # Use dynamic flat rate set by the admin (currently 8.56)
    load_kw = float(getattr(consumer, 'load_kw', 1.0))
    energy_charges = round(units_consumed * rate_per_unit, 2)

    fixed_charges  = round(load_kw * fixed_charge_per_kw, 2)
    duty_charge    = round((energy_charges + fixed_charges) * (duty_percentage / 100), 2)
    meter_rent     = phase_1_rent if consumer.meter_type == 'analog' else phase_3_rent
    arrears        = 0.0
    late_payment_surcharge = 0.0
    regulatory_surcharge   = 0.0

    total_amount = round(
        energy_charges + fixed_charges + duty_charge
        + regulatory_surcharge + meter_rent + arrears + late_payment_surcharge,
        0
    )
    # ─────────────────────────────────────────────────────────────────────────

    return JsonResponse({
        'success': True,
        'consumer_id': consumer.id,
        'consumer_name': consumer.name,
        'meter_type': consumer.meter_type,
        'load_kw': load_kw,
        'previous_reading': previous_reading,
        'current_reading': current_reading,
        'units_consumed': units_consumed,
        'breakdown': {
            'rate_per_unit': rate_per_unit,
            'energy_charges': energy_charges,
            'fixed_charge_per_kw': fixed_charge_per_kw,
            'fixed_charges': fixed_charges,
            'duty_percentage': duty_percentage,
            'duty_charge': duty_charge,
            'meter_rent': meter_rent,
            'regulatory_surcharge': regulatory_surcharge,
            'arrears': arrears,
            'late_payment_surcharge': late_payment_surcharge,
        },
        'total_amount': int(total_amount),
    })

@api_view(['POST'])
@require_role('admin', 'meter_reader')
def api_send_bill_sms(request):
    """Send bill summary SMS to consumer's phone number via Twilio."""
    try:
        data = request.data
        bill_id = data.get('bill_id')

        if not bill_id:
            return JsonResponse({'error': 'bill_id is required'}, status=400)

        bill = get_object_or_404(Bill, id=bill_id)
        consumer = bill.consumer

        if not consumer.phone:
            return JsonResponse({
                'success': False,
                'error': 'Consumer does not have a phone number on file'
            }, status=400)

        # Compose SMS message
        message_body = (
            f"Dear {consumer.name}, your electricity bill has been generated.\n"
            f"Bill#: {bill.bill_number}\n"
            f"Units: {bill.units} kWh\n"
            f"Amount: Rs.{bill.total_amount:.2f}\n"
            f"Due: {bill.due_date.strftime('%d-%b-%Y') if bill.due_date else 'N/A'}\n"
            f"- eMeter AMU"
        )

        # Try sending via Twilio
        twilio_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
        twilio_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
        # Prefer WhatsApp if available/configured, otherwise SMS
        twilio_sender = getattr(settings, 'TWILIO_WHATSAPP_FROM', '') or getattr(settings, 'TWILIO_PHONE_NUMBER', '')

        if not all([twilio_sid, twilio_token, twilio_sender]):
            # Twilio not configured — return the message content for display
            return JsonResponse({
                'success': True,
                'sms_sent': False,
                'reason': 'Twilio credentials not configured. SMS not sent.',
                'message_preview': message_body,
                'phone': consumer.phone,
            })

        try:
            from twilio.rest import Client
            client = Client(twilio_sid, twilio_token)

            # Ensure phone has country code
            phone = consumer.phone.strip()
            if not phone.startswith('+'):
                phone = '+91' + phone.lstrip('0')

            # Use WhatsApp format for the recipient if sending via WhatsApp
            to_phone = phone
            if twilio_sender.startswith('whatsapp:'):
                if not to_phone.startswith('whatsapp:'):
                    to_phone = f'whatsapp:{to_phone}'

            print(f"DEBUG: Attempting to send Twilio message from {twilio_sender} to {to_phone}")
            sms = client.messages.create(
                body=message_body,
                from_=twilio_sender,
                to=to_phone
            )
            print(f"DEBUG: Twilio message sent successfully. SID: {sms.sid}")

            return JsonResponse({
                'success': True,
                'sms_sent': True,
                'message_sid': sms.sid,
                'phone': phone,
                'message_preview': message_body,
            })

        except ImportError:
            return JsonResponse({
                'success': True,
                'sms_sent': False,
                'reason': 'Twilio library not installed. Run: pip install twilio',
                'message_preview': message_body,
                'phone': consumer.phone,
            })
        except Exception as sms_error:
            return JsonResponse({
                'success': True,
                'sms_sent': False,
                'reason': f'SMS send failed: {str(sms_error)}',
                'message_preview': message_body,
                'phone': consumer.phone,
            })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@api_view(['PUT', 'PATCH'])
@require_role('admin', 'meter_reader')
def api_edit_today_reading(request, reading_id):
    """Edit a meter reading — only allowed if reading_date is today."""
    if request.method not in ['PUT', 'PATCH']:
        return JsonResponse({'detail': 'Method not allowed'}, status=405)

    try:
        reading = get_object_or_404(MeterReading, id=reading_id)

        # Check if reading is from today
        today = timezone.localtime().date()
        if reading.reading_date != today:
            return JsonResponse({
                'success': False,
                'error': 'You can only edit readings from today',
                'reading_date': str(reading.reading_date),
                'today': str(today),
            }, status=403)

        data = request.data
        new_current_reading = float(data.get('current_reading', reading.current_reading))

        if new_current_reading < reading.previous_reading:
            return JsonResponse({
                'error': 'Current reading must be >= previous reading',
                'previous_reading': reading.previous_reading,
            }, status=400)

        # Cannot edit reading if the linked bill is locked/finalized
        linked_bill = Bill.objects.filter(meter_reading=reading).first()
        if linked_bill and linked_bill.is_locked:
            return JsonResponse({
                'success': False,
                'error': 'Cannot edit a reading that is already linked to a finalized bill.'
            }, status=400)
            
        # Update the reading
        reading.current_reading = new_current_reading
        reading.units_consumed = new_current_reading - reading.previous_reading
        reading.save()

        return JsonResponse({
            'success': True,
            'reading': {
                'id': reading.id,
                'current_reading': reading.current_reading,
                'units_consumed': reading.units_consumed,
            },
            'bill_updated': False,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@api_view(['GET'])
@require_authenticated
def api_get_bill_pdf(request, bill_id):
    """API endpoint to download branded PDF bill"""
    
    bill = get_object_or_404(Bill, id=bill_id)
    
    # BOLA / IDOR PDF Fetch Fix:
    # Only allow admins, meter readers, or the consumer who owns the bill to download/view the PDF
    if request.user.role not in ['admin', 'meter_reader']:
        if not (bill.consumer and bill.consumer.user == request.user):
            return JsonResponse({'error': 'Permission denied. You do not have access to this bill.'}, status=403)
    
    try:
        # Use snapshot data for PDF if bill is locked
        if bill.is_locked:
            pdf_data = {
                'bill_number': bill.bill_number,
                'bill_date': bill.locked_at.strftime('%d %b %Y') if bill.locked_at else timezone.now().strftime('%d %b %Y'),
                'due_date': bill.due_date.strftime('%d %b %Y') if bill.due_date else 'N/A',
                'connection_type': bill.consumer.connection_type,
                'load_kw': bill.consumer.load_kw,
                'billing_period_start': bill.billing_period_start.strftime('%B %Y') if bill.billing_period_start else 'N/A',
                'consumer_name': bill.consumer_name_snapshot or bill.consumer.name,
                'consumer_number': bill.consumer.consumer_number,
                'meter_number': bill.meter_number_snapshot or bill.consumer.meter_number,
                'address': bill.consumer.address,
                'previous_reading': bill.meter_reading.previous_reading if bill.meter_reading else 0.0,
                'current_reading': bill.meter_reading.current_reading if bill.meter_reading else (bill.units_consumed_snapshot or bill.units),
                'units': bill.units_consumed_snapshot or bill.units,
                'rate_per_unit': bill.rate_snapshot or bill.rate_per_unit,
                'energy_charges': bill.energy_charges,
                'fixed_charges': bill.fixed_charges,
                'duty_charge': bill.duty_charge,
                'meter_rent': bill.meter_rent,
                'meter_type': '1' if bill.consumer.meter_type == 'analog' else '3',
                'regulatory_surcharge': bill.regulatory_surcharge,
                'arrears': bill.arrears,
                'late_payment_surcharge': bill.late_payment_surcharge,
                'total_amount': bill.total_amount_snapshot or bill.total_amount,
                'total_payable': int(round(bill.total_amount_snapshot or bill.total_amount)),
                'current_year': timezone.now().year,
                'is_finalized': True
            }
        else:
            # Fallback for draft bills (limited preview)
            pdf_data = {
                'bill_number': bill.bill_number,
                'bill_date': timezone.now().strftime('%d %b %Y'),
                'due_date': bill.due_date.strftime('%d %b %Y') if bill.due_date else 'N/A',
                'connection_type': bill.consumer.connection_type,
                'load_kw': bill.consumer.load_kw,
                'billing_period_start': bill.billing_period_start.strftime('%B %Y') if bill.billing_period_start else 'N/A',
                'consumer_name': bill.consumer.name,
                'consumer_number': bill.consumer.consumer_number,
                'meter_number': bill.consumer.meter_number,
                'address': bill.consumer.address,
                'previous_reading': bill.meter_reading.previous_reading if bill.meter_reading else 0,
                'current_reading': bill.meter_reading.current_reading if bill.meter_reading else 0,
                'units': bill.units,
                'rate_per_unit': bill.rate_per_unit,
                'energy_charges': bill.energy_charges,
                'fixed_charges': bill.fixed_charges,
                'duty_charge': bill.duty_charge,
                'meter_rent': bill.meter_rent,
                'meter_type': '1' if bill.consumer.meter_type == 'analog' else '3',
                'regulatory_surcharge': bill.regulatory_surcharge,
                'arrears': bill.arrears,
                'late_payment_surcharge': bill.late_payment_surcharge,
                'total_amount': bill.total_amount,
                'total_payable': int(round(bill.total_amount)),
                'current_year': timezone.now().year,
                'is_finalized': False
            }
        
        pdf_content = BillPDFGenerator.generate_bill_pdf(pdf_data)
        
        AuditLog.objects.create(
            user=request.user,
            action='pdf_gen',
            bill=bill,
            details={'is_finalized': bill.is_locked}
        )

        if pdf_content:
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="AMU_Bill_{bill.bill_number}.pdf"'
            return response
        return JsonResponse({'error': 'Failed to generate PDF'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@api_view(['POST'])
@require_role('admin')
def api_manual_generate_bill(request):
    """Manually generate a bill for a consumer with provided data (admin only)"""
        
    try:
        data = request.data
        consumer_id = data.get('consumer_id')
        current_reading = float(data.get('current_reading', 0))
        billing_period_start_str = data.get('billing_period_start')
        due_date_str = data.get('due_date')
        meter_id = data.get('meter_id')
        meter_number = data.get('meter_number')
        created_source = data.get('created_source', 'admin_manual')
        manual_override_reason = data.get('manual_override_reason', '')
            
        with transaction.atomic():
            consumer = get_object_or_404(Consumer.objects.select_for_update(), id=consumer_id)
            
            from billing.models import Meter, ConsumerMeterAssignment
            meter = None
            if meter_id:
                meter = Meter.objects.filter(id=meter_id).first()
            elif meter_number:
                meter = Meter.objects.filter(meter_number=meter_number).first()
            else:
                active_assignments = ConsumerMeterAssignment.objects.filter(
                    consumer=consumer,
                    is_active=True
                )
                if active_assignments.exists():
                    meter = active_assignments.first().meter
                elif consumer.meter_number:
                    meter, _ = Meter.objects.get_or_create(
                        meter_number=consumer.meter_number,
                        defaults={'meter_type': consumer.meter_type}
                    )
            
            if meter:
                # Lock the meter row to prevent concurrent modifications
                meter = Meter.objects.select_for_update().get(id=meter.id)

            # Get last reading per meter
            last_reading = MeterReading.objects.filter(meter=meter).order_by('-reading_date').first() if meter else None
            previous_reading = last_reading.current_reading if last_reading else 0
            
            if current_reading < previous_reading:
                return JsonResponse({'error': f'Current reading cannot be less than previous reading ({previous_reading}) for Meter {meter.meter_number if meter else ""}'}, status=400)
                
            # 1-5 Atomic Generation of Final Bill
            billing_period_start = datetime.strptime(billing_period_start_str + '-01', '%Y-%m-%d').date() if billing_period_start_str else timezone.localtime().date()
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else (timezone.localtime() + timedelta(days=15)).date()
            
            bill = BillingService.generate_final_bill(
                consumer=consumer,
                current_reading=current_reading,
                reading_date=timezone.localtime().date(),
                user=request.user,
                billing_period=billing_period_start,
                due_date=due_date,
                meter=meter,
                created_source=created_source,
                manual_override_reason=manual_override_reason
            )

        return JsonResponse({
            'success': True,
            'bill_id': bill.id,
            'bill_number': bill.bill_number,
            'total_amount': float(bill.total_amount),
            'created_at': bill.created_at.isoformat() if bill.created_at else None,
            'message': 'Bill generated successfully'
        })
    except ValidationError as e:
        return JsonResponse({'error': e.message if hasattr(e, 'message') else str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@api_view(['POST'])
@require_role('admin')
def api_import_readings(request):
    """Import meter readings from an Excel file (admin only).

    Supports two column layouts (auto-detected):

    4-column layout (recommended):
        A: Consumer Number  B: Meter Number  C: Current Reading  D: Reading Date (optional)

    3-column layout (legacy / mobile export):
        A: Consumer Number  B: Current Reading  C: Reading Date (optional)

    Validation steps:
        1. Auto-detect layout from whether column B is numeric or text
        2. Verify Consumer Number exists; if meter number is provided, cross-check it
        3. Check no duplicate reading exists for the same billing month
        4. Validate reading value (must be >= previous), then save + generate bill
    """

    # ← FIXED: Safe role check using user.role
    if request.user.role != 'admin':
        return JsonResponse({'detail': 'Admin authentication required'}, status=403)

    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    excel_file = request.FILES['file']

    # --- DEFENSIVE FILE VALIDATION ---
    # 1. Limit file size to 5MB
    if excel_file.size > 5 * 1024 * 1024:
        return JsonResponse({'error': 'File size exceeds the 5MB limit.'}, status=400)

    # 2. Validate file extension
    file_name = excel_file.name.lower()
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        return JsonResponse({'error': 'Invalid file format. Only Excel files (.xlsx, .xls) are allowed.'}, status=400)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        success_count = 0
        duplicate_count = 0
        error_count = 0
        errors = []
        bills_created = []

        # ── Auto-detect column layout ────────────────────────────────────────
        # Peek at row 2 (first data row).  If col B parses as a number → 3-col;
        # if it looks like text → 4-col.
        layout = 4  # default
        for peek_row in sheet.iter_rows(min_row=2, max_row=3, values_only=True):
            if not peek_row or len(peek_row) < 2:
                break
            col_b = peek_row[1]
            if col_b is not None:
                try:
                    float(col_b)
                    layout = 3  # col B is numeric → reading value
                except (ValueError, TypeError):
                    layout = 4  # col B is text → meter number
            break

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue

            # ── Parse columns based on detected layout ───────────────────────
            if layout == 3:
                # A: Consumer Number, B: Current Reading, C: Reading Date
                if len(row) < 2:
                    continue
                consumer_number  = str(row[0]).strip() if row[0] is not None else None
                meter_number     = None   # not provided — skip cross-check
                current_reading  = row[1]
                reading_date_raw = row[2] if len(row) > 2 else None
            else:
                # A: Consumer Number, B: Meter Number, C: Current Reading, D: Reading Date
                if len(row) < 3:
                    continue
                consumer_number  = str(row[0]).strip() if row[0] is not None else None
                meter_number     = str(row[1]).strip() if row[1] is not None else None
                current_reading  = row[2]
                reading_date_raw = row[3] if len(row) > 3 else None

            # Skip fully empty rows
            if consumer_number is None and current_reading is None:
                continue
            # Skip header-like rows that somehow sneak into data rows
            if consumer_number and str(consumer_number).lower() in ('consumer number', 'consumer_number', 'consumer#'):
                continue

            # ── STEP 1: Verify Consumer Number ──────────────────────────────
            if not consumer_number:
                errors.append(f"Row {row_idx}: Consumer Number is required.")
                error_count += 1
                continue

            try:
                consumer = Consumer.objects.get(consumer_number=consumer_number)
            except Consumer.DoesNotExist:
                errors.append(f"Row {row_idx}: Consumer '{consumer_number}' not found in system.")
                error_count += 1
                continue

            # ── Parse reading date ───────────────────────────────────────────
            if not reading_date_raw:
                reading_date = timezone.localtime().date()
            elif hasattr(reading_date_raw, 'date'):
                reading_date = reading_date_raw.date()
            elif isinstance(reading_date_raw, str):
                reading_date_str = reading_date_raw.strip()
                parsed = None
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        parsed = datetime.strptime(reading_date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                reading_date = parsed if parsed else timezone.localtime().date()
            else:
                reading_date = timezone.localtime().date()

            billing_month = reading_date.replace(day=1)

            # ── Resolve Meter and Verify Assignment ──────────────────────
            from billing.models import Meter, ConsumerMeterAssignment
            from django.db import models
            meter = None
            if meter_number:
                meter = Meter.objects.filter(meter_number=meter_number).first()
                if not meter:
                    errors.append(f"Row {row_idx} ({consumer_number}): Meter '{meter_number}' does not exist in system.")
                    error_count += 1
                    continue
                
                # Check assignment temporally during reading_date
                assignment_exists = ConsumerMeterAssignment.objects.filter(
                    consumer=consumer,
                    meter=meter,
                    start_date__lte=reading_date
                ).filter(
                    models.Q(end_date__gte=reading_date) | models.Q(end_date__isnull=True)
                ).exists()
                
                if not assignment_exists:
                    errors.append(
                        f"Row {row_idx} ({consumer_number}): Meter '{meter_number}' was not assigned to consumer '{consumer_number}' on {reading_date}."
                    )
                    error_count += 1
                    continue
            else:
                # temporal assignment resolve
                active_assignments = ConsumerMeterAssignment.objects.filter(
                    consumer=consumer,
                    start_date__lte=reading_date
                ).filter(
                    models.Q(end_date__gte=reading_date) | models.Q(end_date__isnull=True)
                )
                if active_assignments.exists():
                    meter = active_assignments.first().meter
                elif consumer.meter_number:
                    meter, _ = Meter.objects.get_or_create(
                        meter_number=consumer.meter_number,
                        defaults={'meter_type': consumer.meter_type}
                    )
                else:
                    errors.append(f"Row {row_idx} ({consumer_number}): No active meter assignment found on {reading_date}.")
                    error_count += 1
                    continue

            # ── STEP 2: Duplicate check per Meter ─────────────────────────────
            dup = MeterReading.objects.filter(
                meter=meter,
                reading_date__year=billing_month.year,
                reading_date__month=billing_month.month,
            ).first()
            if dup:
                errors.append(
                    f"Row {row_idx} ({consumer_number}): Reading already exists for "
                    f"{billing_month.strftime('%B %Y')} (on {dup.reading_date}) for Meter '{meter.meter_number}'. Skipped."
                )
                duplicate_count += 1
                continue

            # ── STEP 3: Validate reading value ───────────────────────────────
            try:
                curr_val = float(current_reading)
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx} ({consumer_number}): Invalid reading value '{current_reading}'. Must be a number.")
                error_count += 1
                continue

            last_reading = MeterReading.objects.filter(
                meter=meter,
                reading_date__lt=reading_date,
            ).order_by('-reading_date', '-id').first()

            if not last_reading:
                last_reading = MeterReading.objects.filter(
                    meter=meter
                ).order_by('-reading_date', '-id').first()

            previous_val = float(last_reading.current_reading) if last_reading else float(getattr(consumer, 'initial_reading', 0) or 0)

            if curr_val < previous_val:
                errors.append(
                    f"Row {row_idx} ({consumer_number}): Reading {curr_val} < previous {previous_val} for Meter '{meter.meter_number}'. Rejected."
                )
                error_count += 1
                continue

            units_consumed = curr_val - previous_val

            # ── Save reading + generate bill ─────────────────────────────────
            try:
                with transaction.atomic():
                    due_date = reading_date + timedelta(days=30)
                    bill = BillingService.generate_final_bill(
                        consumer=consumer,
                        current_reading=curr_val,
                        reading_date=reading_date,
                        user=request.user,
                        billing_period=billing_month,
                        due_date=due_date,
                        remarks="Imported from Excel",
                        meter=meter
                    )

                    bills_created.append({
                        'consumer_number': consumer.consumer_number,
                        'consumer_name': consumer.name,
                        'meter_number': meter.meter_number,
                        'bill_number': bill.bill_number,
                        'previous_reading': previous_val,
                        'current_reading': curr_val,
                        'units': round(units_consumed, 2),
                        'total_amount': float(bill.total_amount),
                        'due_date': due_date.strftime('%Y-%m-%d'),
                        'status': bill.status,
                    })
                    success_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx} ({consumer_number}): Failed to save — {str(e)}")
                error_count += 1

        # Create bulk import complete notification
        from billing.models import AdminNotification
        AdminNotification.objects.create(
            title="Import Complete",
            message=f"Import Complete. {success_count} readings imported. {duplicate_count} duplicates skipped.",
            notification_type="success",
            is_read=False
        )

        return JsonResponse({
            'success': True,
            'layout_detected': f'{layout}-column',
            'message': f'Import complete — {success_count} bill(s) generated, {error_count} failed, {duplicate_count} skipped.',
            'success_count': success_count,
            'error_count': error_count,
            'duplicate_count': duplicate_count,
            'bills': bills_created,
            'errors': errors[:50],
        })

    except ImportError:
        return JsonResponse(
            {'error': 'openpyxl not installed. Run: pip install openpyxl'},
            status=500
        )
    except Exception as e:
        import traceback
        return JsonResponse({'error': f'Failed to process file: {str(e)}', 'trace': traceback.format_exc()}, status=500)


def check_consumption_anomaly(reading):
    """
    Computes rolling 3-month average of active billing periods for the consumer.
    If the uploaded reading exceeds the average by > 200%, spawn a Notification error alert.
    """
    try:
        from billing.models import MeterReading, AdminNotification
        
        # Get previous 3 readings for this consumer (excluding the current one)
        prev_readings = MeterReading.objects.filter(
            consumer=reading.consumer,
            reading_date__lt=reading.reading_date
        ).order_by('-reading_date')[:3]
        
        if prev_readings.count() > 0:
            total_units = sum(r.units_consumed for r in prev_readings)
            avg_units = total_units / prev_readings.count()
            
            if avg_units > 0:
                current_units = reading.units_consumed
                increase_ratio = (current_units - avg_units) / avg_units
                if increase_ratio > 2.0:
                    percentage_increase = int(increase_ratio * 100)
                    title = "Consumption Anomaly Alert"
                    message = (
                        f"ANOMALY DETECTED:\n"
                        f"Consumer {reading.consumer.name}\n"
                        f"Meter #{reading.consumer.meter_number}\n"
                        f"Usage increased by {percentage_increase}%.\n"
                        f"Potential leakage or faulty reading."
                    )
                    AdminNotification.objects.create(
                        title=title,
                        message=message,
                        notification_type="error",
                        is_read=False
                    )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error checking consumption anomaly: {e}", exc_info=True)


@api_view(['GET'])
@require_role('admin', 'meter_reader')
def api_get_notifications(request):
    """Retrieve all admin/operator notifications"""
    from billing.models import AdminNotification
    from django.db.models import Q
    
    if request.user.is_authenticated:
        user_filter = Q(user__isnull=True) | Q(user=request.user)
    else:
        user_filter = Q(user__isnull=True)
        
    notifications = AdminNotification.objects.filter(user_filter)[:100]
    unread_count = AdminNotification.objects.filter(user_filter, is_read=False).count()
    
    results = []
    import django.utils.timezone as timezone
    now = timezone.now()
    
    for n in notifications:
        diff = now - n.created_at
        if diff.days > 0:
            time_str = f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
        elif diff.seconds >= 3600:
            hours = diff.seconds // 3600
            time_str = f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif diff.seconds >= 60:
            mins = diff.seconds // 60
            time_str = f"{mins} min{'s' if mins > 1 else ''} ago"
        else:
            time_str = "just now"
            
        results.append({
            'id': str(n.id),
            'title': n.title,
            'description': n.message,
            'time': time_str,
            'type': n.notification_type,
            'read': n.is_read
        })
        
    return JsonResponse({
        'success': True,
        'notifications': results,
        'unread_count': unread_count
    })


@api_view(['POST'])
@require_role('admin', 'meter_reader')
def api_mark_notifications_read(request):
    """Mark all or a specific notification as read"""
    from billing.models import AdminNotification
    from django.db.models import Q
    try:
        if request.user.is_authenticated:
            user_filter = Q(user__isnull=True) | Q(user=request.user)
        else:
            user_filter = Q(user__isnull=True)
            
        data = request.data if request.data else {}
        notification_id = data.get('notification_id')
        if notification_id:
            AdminNotification.objects.filter(user_filter, id=notification_id).update(is_read=True)
        else:
            AdminNotification.objects.filter(user_filter, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
 
 
@api_view(['POST'])
@require_role('admin')
def api_start_reading_cycle(request):
    """
    Start a new monthly reading cycle and assign consumers/zones to readers.
    Creates a notification for all meter readers.
    """
    from billing.models import AdminNotification
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        data = request.data if request.data else {}
        zone = data.get('zone', 'North Zone')
        consumers_count = data.get('consumers_count', 15)
        
        readers = User.objects.filter(role='meter_reader')
        for reader in readers:
            AdminNotification.objects.create(
                user=reader,
                title="New Reading Cycle Started!",
                message=f"New Reading Cycle Started! You have been assigned {consumers_count} consumers in Zone {zone}. Tap to view map/list.",
                notification_type="info"
            )
            
        return JsonResponse({
            'success': True,
            'message': f"Reading cycle started successfully. Notifications sent to {readers.count()} reader(s)."
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ─────────────────────────────────────────────────────────────
# CONSUMER PORTAL API
# New endpoints — zero impact on existing admin/mobile APIs
# ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@require_role('consumer')
def consumer_portal_me(request):
    """Return the logged-in consumer's profile details."""
    try:
        consumer = request.user.consumer_profile
    except Exception:
        return JsonResponse({'detail': 'No consumer profile linked to this account.'}, status=404)

    return JsonResponse({
        'consumer_number': consumer.consumer_number,
        'name': consumer.name,
        'meter_number': consumer.meter_number or '',
        'email': consumer.email or '',
        'phone': consumer.phone or '',
        'address': consumer.address or '',
        'connection_type': consumer.connection_type,
        'billing_type': consumer.billing_type,
        'status': consumer.status,
    })


@api_view(['GET'])
@require_role('consumer')
def consumer_portal_readings(request):
    """Return paginated meter readings for the logged-in consumer."""
    try:
        consumer = request.user.consumer_profile
    except Exception:
        return JsonResponse({'detail': 'No consumer profile linked to this account.'}, status=404)

    readings_qs = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date')

    page_num = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 12))
    paginator = Paginator(readings_qs, page_size)
    page = paginator.get_page(page_num)

    results = [
        {
            'id': r.id,
            'reading_date': r.reading_date.strftime('%Y-%m-%d') if r.reading_date else None,
            'previous_reading': r.previous_reading,
            'current_reading': r.current_reading,
            'units_consumed': r.units_consumed,
            'remarks': r.remarks or '',
        }
        for r in page.object_list
    ]

    return JsonResponse({
        'results': results,
        'count': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page_num,
    })


@api_view(['GET'])
@require_role('consumer')
def consumer_portal_bills(request):
    """Return paginated bills for the logged-in consumer."""
    try:
        consumer = request.user.consumer_profile
    except Exception:
        return JsonResponse({'detail': 'No consumer profile linked to this account.'}, status=404)

    bills_qs = Bill.objects.filter(consumer=consumer).order_by('-bill_date')

    page_num = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 12))
    paginator = Paginator(bills_qs, page_size)
    page = paginator.get_page(page_num)

    results = [
        {
            'id': b.id,
            'bill_number': b.bill_number,
            'created_at': b.created_at.isoformat() if hasattr(b, 'created_at') and b.created_at else None,
            'bill_date': b.bill_date.strftime('%Y-%m-%d') if b.bill_date else None,
            'billing_period_start': b.billing_period_start.strftime('%Y-%m-%d') if b.billing_period_start else None,
            'billing_period_end': b.billing_period_end.strftime('%Y-%m-%d') if b.billing_period_end else None,
            'billing_period': f"{b.billing_period_start.strftime('%b-%y')} to {b.billing_period_end.strftime('%b-%y')}" if b.billing_period_start and b.billing_period_end else None,
            'due_date': b.due_date.strftime('%Y-%m-%d') if b.due_date else None,
            'connection_type': b.consumer.connection_type,
            'billing_type': b.consumer.billing_type,
            'load_kw': float(b.consumer.load_kw),
            'meter_type': b.consumer.meter_type,
            'consumer_name': b.consumer.name,
            'consumer_number': b.consumer.consumer_number,
            'meter_number': b.consumer.meter_number,
            'address': b.consumer.address,
            'previous_reading': b.meter_reading.previous_reading if hasattr(b, 'meter_reading') and b.meter_reading else 0,
            'current_reading': b.meter_reading.current_reading if hasattr(b, 'meter_reading') and b.meter_reading else 0,
            'units': b.units,
            'rate_per_unit': float(b.rate_per_unit),
            'energy_charges': float(b.energy_charges),
            'fixed_charges': float(b.fixed_charges),
            'duty_charge': float(b.duty_charge),
            'meter_rent': float(b.meter_rent),
            'regulatory_surcharge': float(b.regulatory_surcharge),
            'arrears': float(b.arrears),
            'late_payment_surcharge': float(b.late_payment_surcharge),
            'total_amount': float(b.total_amount),
            'paid_amount': float(b.paid_amount),
            'status': b.status,
            'is_paid': b.status == 'paid',
        }
        for b in page.object_list
    ]

    return JsonResponse({
        'results': results,
        'count': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page_num,
    })


@api_view(['POST'])
@require_role('consumer')
def consumer_portal_change_password(request):
    """Allow a consumer to change their own password."""
    old_password = request.data.get('old_password', '').strip()
    new_password = request.data.get('new_password', '').strip()
    confirm_password = request.data.get('confirm_password', '').strip()

    if not old_password or not new_password or not confirm_password:
        return JsonResponse({'detail': 'All fields are required.'}, status=400)

    if new_password != confirm_password:
        return JsonResponse({'detail': 'New password and confirmation do not match.'}, status=400)

    if len(new_password) < 6:
        return JsonResponse({'detail': 'Password must be at least 6 characters.'}, status=400)

    if not request.user.check_password(old_password):
        return JsonResponse({'detail': 'Current password is incorrect.'}, status=400)

    request.user.set_password(new_password)
    request.user.save()

    # Invalidate existing token so user must re-login after password change
    Token.objects.filter(user=request.user).delete()

    return JsonResponse({'success': True, 'message': 'Password changed successfully. Please log in again.'})


@api_view(['POST'])
@require_role('admin')
def admin_reset_consumer_password(request, consumer_id):
    """
    Admin-only: reset a consumer's portal password back to their consumer_number.
    If the consumer doesn't have a portal account yet, it creates one.
    """
    try:
        consumer = Consumer.objects.get(id=consumer_id)
    except Consumer.DoesNotExist:
        return JsonResponse({'detail': 'Consumer not found.'}, status=404)

    if not consumer.user:
        # Auto-create the user account if it doesn't exist yet
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            user = User.objects.create_user(
                username=consumer.consumer_number,
                password=consumer.consumer_number,
                role='consumer',
                first_name=consumer.name.split()[0] if consumer.name else '',
                last_name=' '.join(consumer.name.split()[1:]) if consumer.name and len(consumer.name.split()) > 1 else '',
                email=consumer.email or '',
            )
            consumer.user = user
            consumer.save(update_fields=['user'])
        except Exception as e:
            return JsonResponse({'detail': f'Failed to create portal account: {str(e)}'}, status=400)
    else:
        # Reset password for existing user
        consumer.user.set_password(consumer.consumer_number)
        consumer.user.save()

        # Invalidate any existing tokens so the consumer must re-login
        from rest_framework.authtoken.models import Token
        Token.objects.filter(user=consumer.user).delete()

    return JsonResponse({
        'success': True,
        'message': f"Password for {consumer.name} reset to their consumer number ({consumer.consumer_number})."
    })


@api_view(['POST'])
@require_role('admin')
def admin_create_consumer_portal_account(request, consumer_id):
    """
    Admin-only: create a portal User account for an existing consumer.
    Username = consumer_number, default password = consumer_number.
    Idempotent — safe to call even if account already exists.
    """
    try:
        consumer = Consumer.objects.get(id=consumer_id)
    except Consumer.DoesNotExist:
        return JsonResponse({'detail': 'Consumer not found.'}, status=404)

    if consumer.user:
        return JsonResponse({
            'success': True,
            'already_exists': True,
            'message': f'Portal account already exists for {consumer.name} (username: {consumer.user.username}).'
        })

    # Create the User account
    user = User.objects.create_user(
        username=consumer.consumer_number,
        password=consumer.consumer_number,
        role=User.Role.CONSUMER,
        first_name=consumer.name.split()[0] if consumer.name else '',
        last_name=' '.join(consumer.name.split()[1:]) if consumer.name and len(consumer.name.split()) > 1 else '',
        email=consumer.email or '',
    )
    consumer.user = user
    consumer.save(update_fields=['user'])

    return JsonResponse({
        'success': True,
        'already_exists': False,
        'message': f'Portal account created for {consumer.name}. Username: {consumer.consumer_number}.'
    })


@api_view(['GET'])
@require_role('admin')
def export_mobile_sync(request):
    """
    Exports a JSON file containing all data needed for the mobile app
    to operate in 'Total Offline Mode'.
    """
    import json
    from datetime import datetime
    from django.http import HttpResponse


    # 2. Get Consumers
    consumers = Consumer.objects.filter(status='active')
    consumers_data = []
    for c in consumers:
        last_reading = MeterReading.objects.filter(consumer=c).order_by('-reading_date', '-id').first()
        prev_val = last_reading.current_reading if last_reading else 0.0

        consumers_data.append({
            'id': c.id,
            'consumer_number': c.consumer_number,
            'name': c.name,
            'email': c.email,
            'phone': c.phone,
            'address': c.address,
            'meter_number': c.meter_number,
            'meter_type': c.connection_type,
            'load_kw': float(c.load_kw) if c.load_kw else 1.0,
            'previous_reading': float(prev_val),
            'status': c.status
        })
        
    # 3. Get Billing Settings
    try:
        settings = BillingSettings.objects.get(id=1)
        settings_data = {
            'rate_per_unit': float(settings.rate_per_unit),
            'fixed_charge_per_kw': float(settings.fixed_charge_per_kw),
            'duty_percentage': float(settings.duty_percentage),
            'phase_1_rent': float(settings.phase_1_rent),
            'phase_3_rent': float(settings.phase_3_rent),
        }
    except BillingSettings.DoesNotExist:
        # Defaults if settings not configured
        settings_data = {
            'rate_per_unit': 6.50,
            'fixed_charge_per_kw': 50.00,
            'duty_percentage': 5.00,
            'phase_1_rent': 10.00,
            'phase_3_rent': 25.00,
        }
        
    sync_data = {
        'version': 1,
        'exported_at': datetime.now().isoformat(),
        'consumers': consumers_data,
        'settings': settings_data
    }
    
    response = HttpResponse(
        json.dumps(sync_data),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="eMeter_Sync_{datetime.now().strftime("%Y%m%d_%H%M")}.json"'
    return response